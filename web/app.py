import os
import re
import sys
import json
from datetime import datetime
from io import StringIO, BytesIO
import xml.etree.ElementTree as ET

import pandas as pd
from flask import Flask, render_template, request, jsonify, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# ============================================================
# RUTAS BASE Y CONFIGURACIÓN INICIAL
# ============================================================
BASE_PATH = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
MODULOS_PATH = os.path.join(BASE_PATH, "modulos")
EXCEL_PATH = os.path.join(BASE_PATH, "Excel")
EXPORT_PATH = os.path.join(BASE_PATH, "Export")
LOG_PATH = os.path.join(BASE_PATH, "logs")

os.makedirs(LOG_PATH, exist_ok=True)
os.makedirs(EXCEL_PATH, exist_ok=True)
os.makedirs(EXPORT_PATH, exist_ok=True)
# Limpieza automática al iniciar el servidor
for carpeta in [EXCEL_PATH, EXPORT_PATH]:
    try:
        for archivo in os.listdir(carpeta):
            ruta = os.path.join(carpeta, archivo)
            if os.path.isfile(ruta):
                os.remove(ruta)
    except Exception:
        pass

if MODULOS_PATH not in sys.path:
    sys.path.append(MODULOS_PATH)

# ============================================================
# IMPORTACIÓN DE MÓDULOS DEL PROYECTO
# ============================================================
from validador import (
    leer_hoja_eventos,
    leer_hoja_completar,
    leer_hoja_imput,
    ejecutar_flujo_tradicional,
    validar_promocion_tradicional,
    validar_promocion_completar,
    normalizar_local,
    normalizar_texto,
    parsear_promos,
    convertir_txt_a_xml_con_root,
)

try:
    from parser_listas_export import parsear_listas_productos_export
except Exception:
    def parsear_listas_productos_export(_ruta):
        return {}

try:
    from gestor import registrar_rutas_gestor
except Exception:
    def registrar_rutas_gestor(_app):
        return None

try:
    from repositorio import registrar_rutas_repositorio
except Exception:
    def registrar_rutas_repositorio(_app):
        return None

# ============================================================
# CONFIGURACIÓN FLASK
# ============================================================
app = Flask(
    __name__,
    template_folder=os.path.join(os.path.dirname(__file__), "templates"),
    static_folder=os.path.join(os.path.dirname(__file__), "static"),
)
app.secret_key = "ClaveUltraSecretaParaMensajesWeb"
registrar_rutas_gestor(app)
registrar_rutas_repositorio(app)

ULTIMO_REPORTE_DESCARGA = {
    "rc": "",
    "tradicional": [],
    "completar": [],
}

# ============================================================
# LOGGING
# ============================================================
def escribir_log(linea):
    archivo = os.path.join(LOG_PATH, f"log_{datetime.now().strftime('%Y-%m-%d')}.txt")
    with open(archivo, "a", encoding="utf-8") as f:
        f.write(f"[{datetime.now().strftime('%H:%M:%S')}] {linea}\n")


# ============================================================
# UTILIDADES GENERALES DE ARCHIVOS
# ============================================================
def limpiar_carpeta(path):
    errores = []
    for f in os.listdir(path):
        try:
            fp = os.path.join(path, f)
            if os.path.isfile(fp):
                os.remove(fp)
        except Exception as e:
            errores.append(str(e))
    return errores


def listar_archivos():
    return sorted(os.listdir(EXCEL_PATH)), sorted(os.listdir(EXPORT_PATH))


# ============================================================
# UTILIDADES DE LIMPIEZA Y FORMATEO
# ============================================================
def _strip_html(texto):
    return re.sub(r"<[^>]+>", "", str(texto or "")).strip()


def _extraer_entre_parentesis(texto, etiqueta):
    patron = rf"{re.escape(etiqueta)}\s*:?\s*\((.*?)\)"
    m = re.search(patron, texto)
    return m.group(1).strip() if m else ""


def _normalizar_lista_valores(valor):
    if not valor or valor == "-":
        return "-"
    partes = [p.strip() for p in valor.split(",") if p.strip()]
    return " - ".join(partes) if partes else "-"


def _formatear_monto_limpio(valor):
    if not valor or valor == "-":
        return "-"
    try:
        num = float(str(valor).replace(",", "."))
        return f"${int(num)}" if num.is_integer() else f"${num:.2f}"
    except Exception:
        return str(valor)


def _formatear_numero_limpio(valor):
    if not valor or valor == "-":
        return "-"
    try:
        num = float(str(valor).replace(",", "."))
        return f"{num:.2f}"
    except Exception:
        return str(valor)


def _formatear_porcentaje_limpio(valor):
    return "-" if not valor or valor == "-" else valor


# ============================================================
# ANÁLISIS DE DETALLES PARA RESUMEN WEB
# ============================================================
def analizar_detalles(detalles):
    mensajes = []
    for d in detalles:
        if isinstance(d, tuple):
            tipo, msg = d
        else:
            tipo, msg = d.get("tipo"), d.get("msg")
        mensajes.append({"tipo": tipo, "msg": msg, "msg_plain": _strip_html(msg)})

    resumen = {
        "estado_id": "No evaluado",
        "area_responsable": "-",
        "estado_facturar": "No evaluado",
        "estado_fechas": "No evaluado",
        "estado_condicion": "No evaluado",
        "estado_applier": "No evaluado",
        "fecha_inicio_ok": None,
        "fecha_fin_ok": None,
        "tipo_promocion": "-",
        "resumen_condicion": "-",
        "resumen_aplicador": "-",
        "mensaje_principal": "No coinciden",
        "aviso_principal": "",
    }

    id_items = [x for x in mensajes if x["msg_plain"].startswith("[ID]")]
    fact_items = [x for x in mensajes if x["msg_plain"].startswith("[FACTURAR]")]
    fechas_items = [x for x in mensajes if x["msg_plain"].startswith("[FECHAS]")]
    condicion_items = [x for x in mensajes if x["msg_plain"].startswith("[CONDICIÓN]")]
    applier_items = [x for x in mensajes if x["msg_plain"].startswith("[APPLIER]")]
    leyenda_items = [x for x in mensajes if x["msg_plain"].startswith("[LEYENDA]")]
    area_items = [x for x in mensajes if x["msg_plain"].startswith("[ÁREA]")]
    descuento_items = [x for x in mensajes if x["msg_plain"].startswith("[DESCUENTO]")]
    lista_items = [x for x in mensajes if x["msg_plain"].startswith("[LISTA PRODUCTOS]")]
    msje_items = [x for x in mensajes if x["msg_plain"].startswith("[MSJE]")]

    if area_items:
        m_area = re.search(r"AreaResponsable detectada:\s*\((.*?)\)", area_items[0]["msg_plain"], re.IGNORECASE)
        if m_area:
            resumen["area_responsable"] = m_area.group(1).strip()

    if id_items:
        resumen["estado_id"] = "Coinciden" if all(x["tipo"] == "OK" for x in id_items) else "No coinciden"

    if fact_items:
        if any(x["tipo"] == "ERR" for x in fact_items):
            resumen["estado_facturar"] = "No coinciden"
        elif any(x["tipo"] == "WARN" for x in fact_items):
            resumen["estado_facturar"] = "Advertencia"
        else:
            resumen["estado_facturar"] = "Coinciden"

    inicio_item = next((x for x in fechas_items if x["msg_plain"].startswith("[FECHAS] Fecha Inicio Excel")), None)
    fin_item = next((x for x in fechas_items if x["msg_plain"].startswith("[FECHAS] Fecha Fin Excel")), None)
    inicio_tipo = inicio_item["tipo"] if inicio_item else None
    fin_tipo = fin_item["tipo"] if fin_item else None

    fecha_inicio_excel = ""
    fecha_fin_excel = ""
    if inicio_item:
        resumen["fecha_inicio_ok"] = (inicio_tipo == "OK")
        m = re.search(r"Fecha Inicio Excel \((.*?)\).*?Export \((.*?)\)", inicio_item["msg_plain"], re.IGNORECASE)
        if m:
            fecha_inicio_excel = m.group(1).strip()
    if fin_item:
        resumen["fecha_fin_ok"] = (fin_tipo == "OK")
        m = re.search(r"Fecha Fin Excel \((.*?)\).*?Export \((.*?)\)", fin_item["msg_plain"], re.IGNORECASE)
        if m:
            fecha_fin_excel = m.group(1).strip()

    estado_fechas_base = "No evaluado"
    if inicio_tipo == "OK" and fin_tipo == "OK":
        estado_fechas_base = "OK"
    elif inicio_tipo in {"WARN", "ERR"} and fin_tipo == "OK":
        estado_fechas_base = "Posible Extensión"
    elif fin_tipo == "ERR":
        estado_fechas_base = "No coinciden"
    elif inicio_item or fin_item:
        estado_fechas_base = "No coinciden"

    detalle_fechas = []
    if fecha_inicio_excel:
        detalle_fechas.append(f"Inicio: {fecha_inicio_excel}")
    if fecha_fin_excel:
        detalle_fechas.append(f"Fin: {fecha_fin_excel}")
    resumen["estado_fechas"] = f"{estado_fechas_base} | {' | '.join(detalle_fechas)}" if detalle_fechas and estado_fechas_base != "No evaluado" else estado_fechas_base

    if condicion_items:
        if any(x["tipo"] == "ERR" for x in condicion_items):
            resumen["estado_condicion"] = "No coinciden"
        elif any(x["tipo"] == "WARN" for x in condicion_items):
            resumen["estado_condicion"] = "Advertencia"
        else:
            resumen["estado_condicion"] = "Coinciden"

    applier_sin_sku_explicito = any("no informa sku explícitos" in x["msg_plain"].lower() or "no informa sku explicitos" in x["msg_plain"].lower() for x in applier_items)
    if applier_items:
        if applier_sin_sku_explicito or any(x["tipo"] == "ERR" for x in applier_items):
            resumen["estado_applier"] = "No coinciden"
        elif any(x["tipo"] == "WARN" for x in applier_items):
            resumen["estado_applier"] = "Advertencia"
        else:
            resumen["estado_applier"] = "Coinciden"

    leyenda_excel = next((x for x in leyenda_items if "Excel → Tipo:" in x["msg_plain"]), None)
    leyenda_cond = next((x for x in leyenda_items if "Condición Export →" in x["msg_plain"]), None)
    leyenda_applier = next((x for x in leyenda_items if "Applier Export →" in x["msg_plain"]), None)

    if leyenda_excel:
        tipo = _extraer_entre_parentesis(leyenda_excel["msg_plain"], "Tipo")
        resumen["tipo_promocion"] = tipo if tipo else "-"

    tipo_prom = (resumen["tipo_promocion"] or "").upper()
    es_msje = "MSJE" in tipo_prom or any("MSJE / POPUP" in x["msg_plain"] for x in msje_items)
    es_2da = "2DA" in tipo_prom
    es_pack = bool(re.search(r"\bPACK\b", tipo_prom) or re.search(r"\d+\s*X\s*\d+", tipo_prom))
    es_porcentual = ("PORCENT" in tipo_prom or "%" in tipo_prom)
    es_nominal = ("NOMINAL" in tipo_prom and "PACK NOMINAL" not in tipo_prom)

    if leyenda_cond:
        sku_val = _extraer_entre_parentesis(leyenda_cond["msg_plain"], "SKU")
        lista_val = _extraer_entre_parentesis(leyenda_cond["msg_plain"], "Lista")
        cantidad_cond_val = _extraer_entre_parentesis(leyenda_cond["msg_plain"], "Cantidad")
        sku_fmt = _normalizar_lista_valores(sku_val)
        lista_fmt = lista_val if lista_val and lista_val != "-" else "-"
        partes_cond = []
        if sku_fmt != "-":
            partes_cond.append(f"SKU: {sku_fmt}")
        elif lista_fmt != "-":
            partes_cond.append(f"Lista: {lista_fmt}")
        if es_2da and cantidad_cond_val and cantidad_cond_val not in {"-", "0", "0.0", "0.00"}:
            try:
                q = float(cantidad_cond_val)
                partes_cond.append(f"Cada {int(q) if q.is_integer() else q} unidades")
            except Exception:
                partes_cond.append(f"Cada {cantidad_cond_val} unidades")
        resumen["resumen_condicion"] = " | ".join(partes_cond) if partes_cond else "-"

    if leyenda_applier:
        sku_val = _extraer_entre_parentesis(leyenda_applier["msg_plain"], "SKU")
        lista_val = _extraer_entre_parentesis(leyenda_applier["msg_plain"], "Lista")
        cantidad_val = _extraer_entre_parentesis(leyenda_applier["msg_plain"], "Cantidad")
        porcentaje_val = _extraer_entre_parentesis(leyenda_applier["msg_plain"], "%")
        monto_val = _extraer_entre_parentesis(leyenda_applier["msg_plain"], "Monto")
        monto_export_val = _extraer_entre_parentesis(leyenda_applier["msg_plain"], "Monto export")
        pct_nodo_export = _extraer_entre_parentesis(leyenda_applier["msg_plain"], "% nodo export")
        pct_comercial_excel = _extraer_entre_parentesis(leyenda_applier["msg_plain"], "% comercial Excel")
        dcto_bruto_q = _extraer_entre_parentesis(leyenda_applier["msg_plain"], "Dcto bruto Excel(Q)")
        pvp_pack_excel = _extraer_entre_parentesis(leyenda_applier["msg_plain"], "PVPOfertaPack Excel")
        if not pvp_pack_excel and leyenda_excel:
            pvp_pack_excel = _extraer_entre_parentesis(leyenda_excel["msg_plain"], "PVPOfertaPack")
        unidades_excel = _extraer_entre_parentesis(leyenda_excel["msg_plain"], "Unidades") if leyenda_excel else ""
        sku_fmt = _normalizar_lista_valores(sku_val)
        lista_fmt = _normalizar_lista_valores(lista_val)
        partes = []
        if sku_fmt != "-":
            partes.append(f"SKU: {sku_fmt}")
        elif lista_fmt != "-":
            partes.append(f"Lista: {lista_fmt}")
        if es_msje:
            salida = _extraer_entre_parentesis(leyenda_applier["msg_plain"], "Salida")
            mensaje = _extraer_entre_parentesis(leyenda_applier["msg_plain"], "Mensaje")
            texto = _extraer_entre_parentesis(leyenda_applier["msg_plain"], "Texto")
            if mensaje and mensaje != "-":
                partes.append(f"Mensaje: {mensaje}")
            if salida and salida != "-":
                partes.append(f"Salida: {salida}")
            if texto and texto != "-":
                partes.append(f"Texto: {texto}")
        elif es_2da:
            if cantidad_val and cantidad_val not in {"-", "0", "0.0", "0.00"}:
                try:
                    q = float(cantidad_val)
                    partes.append(f"Cada {int(q) if q.is_integer() else q} unidades")
                except Exception:
                    partes.append(f"Cada {cantidad_val} unidades")
            pct_aplicador_fmt = _formatear_porcentaje_limpio(pct_nodo_export)
            pct_comercial_fmt = _formatear_porcentaje_limpio(pct_comercial_excel)
            dcto_bruto_fmt = _formatear_numero_limpio(dcto_bruto_q)
            if pct_aplicador_fmt != "-":
                partes.append(f"{pct_aplicador_fmt} aplicador")
            if pct_comercial_fmt != "-":
                partes.append(f"{pct_comercial_fmt} comercial")
            if dcto_bruto_fmt != "-":
                partes.append(f"Dcto bruto Q: {dcto_bruto_fmt}")
        elif es_pack:
            monto_fmt = _formatear_monto_limpio(monto_val or monto_export_val)
            pvp_fmt = _formatear_monto_limpio(pvp_pack_excel)
            if cantidad_val and cantidad_val not in {"-", "0", "0.0", "0.00"}:
                try:
                    q = float(cantidad_val)
                    partes.append(f"Cantidad: {int(q) if q.is_integer() else q}")
                except Exception:
                    partes.append(f"Cantidad: {cantidad_val}")
            if monto_fmt != "-":
                partes.append(f"Monto unitario: {monto_fmt}")
            if pvp_fmt != "-" and unidades_excel and unidades_excel != "-":
                try:
                    q_pack = float(unidades_excel)
                    q_txt = int(q_pack) if q_pack.is_integer() else q_pack
                except Exception:
                    q_txt = unidades_excel
                partes.append(f"Pack: {pvp_fmt} / {q_txt}")
        elif es_nominal:
            pvp_fmt = _formatear_monto_limpio(pvp_pack_excel)
            monto_export_fmt = _formatear_monto_limpio(monto_export_val or monto_val)
            dcto_bruto_fmt = _formatear_numero_limpio(dcto_bruto_q)
            if pvp_fmt != "-":
                partes.append(f"PVPOfertaPack: {pvp_fmt}")
            if monto_export_fmt != "-":
                partes.append(f"Monto export: {monto_export_fmt}")
            if dcto_bruto_fmt != "-":
                partes.append(f"Dcto bruto Q: {dcto_bruto_fmt}")
        elif es_porcentual:
            pct_fmt = _formatear_porcentaje_limpio(porcentaje_val or pct_nodo_export)
            if pct_fmt != "-":
                partes.append(f"%: {pct_fmt}")
        else:
            if cantidad_val and cantidad_val not in {"-", "0", "0.0", "0.00"}:
                try:
                    q = float(cantidad_val)
                    partes.append(f"Cantidad: {int(q) if q.is_integer() else q}")
                except Exception:
                    partes.append(f"Cantidad: {cantidad_val}")
            monto_fmt = _formatear_monto_limpio(monto_export_val or monto_val)
            if monto_fmt != "-":
                partes.append(f"Monto: {monto_fmt}")
        resumen["resumen_aplicador"] = " | ".join(partes) if partes else "-"

    if applier_sin_sku_explicito:
        resumen["resumen_aplicador"] = "ERROR: applier sin SKU explícito"

    if resumen["tipo_promocion"] == "-":
        for x in descuento_items:
            txt = x["msg_plain"]
            m_ambos = re.search(r"Excel\s*\((.*?)\).*?Export\s*\((.*?)\)", txt, re.IGNORECASE)
            if m_ambos:
                resumen["tipo_promocion"] = f"PORCENTUAL - {m_ambos.group(1).strip()}"
                break

    if resumen["resumen_condicion"] == "-":
        for x in condicion_items:
            txt = x["msg_plain"]
            m_lista = re.search(r"misma lista de productos del Excel:\s*\((.*?)\)", txt, re.IGNORECASE)
            if m_lista:
                resumen["resumen_condicion"] = f"Lista: {m_lista.group(1).strip()}"
                break
        if resumen["resumen_condicion"] == "-":
            for x in lista_items:
                txt = x["msg_plain"]
                m = re.search(r"LISTA PRODUCTOS Excel\s*\((.*?)\)", txt, re.IGNORECASE)
                if m:
                    resumen["resumen_condicion"] = f"Lista: {m.group(1).strip()}"
                    break

    hay_err_id = any(x["tipo"] == "ERR" for x in id_items)
    hay_err_fact = any(x["tipo"] == "ERR" for x in fact_items)
    hay_err_cond = any(x["tipo"] == "ERR" for x in condicion_items)
    hay_err_applier = applier_sin_sku_explicito or any(x["tipo"] == "ERR" for x in applier_items)
    solo_ext_fecha_inicio = inicio_tipo == "WARN" and fin_tipo == "OK" and not hay_err_id and not hay_err_fact and not hay_err_cond and not hay_err_applier

    if solo_ext_fecha_inicio:
        resumen["mensaje_principal"] = "Coinciden"
        resumen["aviso_principal"] = "Posible extensión: fecha inicio diferente"
    else:
        if hay_err_id or hay_err_fact or hay_err_cond or hay_err_applier or fin_tipo == "ERR" or (inicio_tipo == "ERR" and fin_tipo != "OK"):
            resumen["mensaje_principal"] = "No coinciden"
        elif inicio_tipo == "OK" and fin_tipo == "OK":
            resumen["mensaje_principal"] = "Coinciden"
        else:
            resumen["mensaje_principal"] = "No coinciden"

    return resumen


def extraer_msje_popup_desde_detalles(detalles):
    resultado = {
        "hay": False,
        "id_msje": "",
        "id_padre": "",
        "mensaje": "No hay",
        "salida": "-",
        "texto": "-",
        "resumen_condicion": "-",
        "resumen_aplicador": "-",
        "detalle": [],
    }

    mensajes = []
    for d in detalles:
        if isinstance(d, tuple):
            tipo, msg = d
        else:
            tipo, msg = d.get("tipo"), d.get("msg")
        plain = _strip_html(msg)
        if plain.startswith("[MSJE]") or plain.startswith("[CONDICIÓN]") or plain.startswith("[APPLIER]"):
            mensajes.append({"tipo": tipo, "msg": msg, "plain": plain})

    rel = next((x for x in mensajes if x["plain"].startswith("[MSJE]") and "corresponde a MSJE / POPUP asociado a promoción" in x["plain"]), None)
    if not rel:
        return resultado

    m = re.search(r"ID Lista Locales\s*\((.*?)\).*?promoción\s*\((.*?)\)", rel["plain"], re.IGNORECASE)
    if not m:
        return resultado

    resultado["hay"] = True
    resultado["id_msje"] = m.group(1).strip()
    resultado["id_padre"] = m.group(2).strip()

    cond = next((x for x in mensajes if x["plain"].startswith("[CONDICIÓN]") and ("condición del MSJE" in x["plain"].lower() or "sku de la condición del msje" in x["plain"].lower() or "Lista de condición del MSJE" in x["plain"])), None)
    apl = next((x for x in mensajes if x["plain"].startswith("[MSJE]") and "Texto del mensaje" in x["plain"]), None)
    salida = next((x for x in mensajes if x["plain"].startswith("[MSJE]") and "Salida del mensaje correcta" in x["plain"]), None)
    resumen = next((x for x in mensajes if x["plain"].startswith("[MSJE]") and "Tipo (MSJE)" in x["plain"]), None)

    if cond:
        resultado["resumen_condicion"] = cond["plain"].replace("[CONDICIÓN]", "").strip()
    if salida:
        ms = re.search(r"\((.*?)\)", salida["plain"])
        if ms:
            resultado["salida"] = ms.group(1).strip()
    if apl:
        mt = re.search(r"Texto del mensaje:\s*\((.*?)\)", apl["plain"], re.IGNORECASE)
        if mt:
            resultado["texto"] = mt.group(1).strip()
    if resumen:
        resultado["resumen_aplicador"] = resumen["plain"].replace("[MSJE]", "").strip()

    resultado["mensaje"] = f"MSJE / POPUP asociado a promoción #{resultado['id_padre']}"
    resultado["detalle"] = [{"tipo": x["tipo"], "msg": x["msg"]} for x in mensajes]
    return resultado


def construir_resultado_web(id_geo, excel_origen, export_origen, promo_info, detalles, analisis, es_msje_popup=False, id_padre="", msje_popup=None):
    msje_popup = msje_popup or {}
    msje_popup_hay = bool(msje_popup.get("hay"))
    msje_popup_id = str(msje_popup.get("id_msje") or "")
    msje_popup_id_padre = str(msje_popup.get("id_padre") or id_geo or "")
    msje_popup_mensaje = msje_popup.get("mensaje") or ("No hay" if not msje_popup_hay else "-")
    msje_popup_salida = msje_popup.get("salida") or "-"
    msje_popup_texto = msje_popup.get("texto") or "-"
    msje_popup_resumen = msje_popup.get("resumen_aplicador") or msje_popup.get("resumen") or "-"
    msje_popup_condicion = msje_popup.get("resumen_condicion") or "-"
    msje_popup_fecha_inicio = msje_popup.get("fecha_inicio") or ""
    msje_popup_fecha_fin = msje_popup.get("fecha_fin") or ""

    busqueda_ids = " ".join(
        p for p in [str(id_geo or "").strip(), str(id_padre or "").strip(), msje_popup_id] if p
    ).strip()

    return {
        "id_geo": str(id_geo),
        "mensaje": analisis["mensaje_principal"],
        "aviso_principal": analisis["aviso_principal"],
        "excel_origen": excel_origen,
        "export_origen": export_origen,
        "promo_info": promo_info,
        "detalle": [{"tipo": d[0], "msg": d[1]} if isinstance(d, tuple) else d for d in detalles],
        "estado_id": analisis["estado_id"],
        "estado_facturar": analisis["estado_facturar"],
        "estado_fechas": analisis["estado_fechas"],
        "estado_condicion": analisis["estado_condicion"],
        "estado_applier": analisis["estado_applier"],
        "fecha_inicio_ok": analisis["fecha_inicio_ok"],
        "fecha_fin_ok": analisis["fecha_fin_ok"],
        "tipo_promocion": analisis["tipo_promocion"],
        "resumen_condicion": analisis["resumen_condicion"],
        "resumen_aplicador": analisis["resumen_aplicador"],
        "es_msje_popup": es_msje_popup,
        "id_padre": str(id_padre or ""),
        "msje_popup_hay": msje_popup_hay,
        "msje_popup_id": msje_popup_id,
        "msje_popup_id_padre": msje_popup_id_padre,
        "msje_popup_mensaje": msje_popup_mensaje,
        "msje_popup_salida": msje_popup_salida,
        "msje_popup_texto": msje_popup_texto,
        "msje_popup_resumen": msje_popup_resumen,
        "msje_popup_condicion": msje_popup_condicion,
        "msje_popup_fecha_inicio": msje_popup_fecha_inicio,
        "msje_popup_fecha_fin": msje_popup_fecha_fin,
        "busqueda_ids": busqueda_ids,
    }




def _copiar_resultados_para_descarga(resultados):
    limpios = []
    for r in resultados:
        copia = dict(r)
        copia["promo_info"] = dict(copia.get("promo_info") or {})
        copia["detalle"] = list(copia.get("detalle") or [])
        limpios.append(copia)
    return limpios


def _estado_reporte_descarga(resultado):
    mensaje = str(resultado.get("mensaje") or "").strip()
    aviso = str(resultado.get("aviso_principal") or "").strip()
    if mensaje == "Coinciden" and aviso:
        return "ATENCION"
    if mensaje == "Coinciden":
        return "OK"
    return "ERROR"


def _valor_si_no(valor):
    return "Sí" if bool(valor) else "No"


def _texto_competencia_prolijo(valor):
    txt = str(valor or "-").strip()
    if not txt or txt == "-":
        return "-"
    txt = txt.replace("Comp. X Producto", "Comp. Por Producto")
    txt = txt.replace("Comp. X Promoción", "Comp. Por Promoción")
    txt = txt.replace("Comp. X Promocion", "Comp. Por Promoción")
    txt = txt.replace("Comp. X Unidades", "Comp. Por Unidades")
    return txt


def _extraer_productos_y_detalle(texto):
    base = str(texto or "-").strip()
    if not base or base == "-":
        return "-", "-"

    partes = [p.strip() for p in base.split("|") if p.strip()]
    if not partes:
        return "-", "-"

    producto = "-"
    detalle = []

    for parte in partes:
        lower = parte.lower()
        if lower.startswith("sku:") or lower.startswith("lista:"):
            if producto == "-":
                producto = parte
            else:
                detalle.append(parte)
        else:
            detalle.append(parte)

    detalle_txt = " | ".join(detalle) if detalle else "-"
    return producto, detalle_txt


def _observacion_reporte(resultado):
    estado = _estado_reporte_descarga(resultado)
    aviso = str(resultado.get("aviso_principal") or "").strip()
    if estado == "OK":
        return "-"
    if aviso:
        return aviso

    detalle = resultado.get("detalle") or []
    preferidos = []
    for d in detalle:
        tipo = str((d or {}).get("tipo") or "").strip().upper()
        msg = _strip_html((d or {}).get("msg", ""))
        if not msg:
            continue
        if tipo == "ERR":
            preferidos.append(msg)
    if preferidos:
        return preferidos[0]

    for d in detalle:
        tipo = str((d or {}).get("tipo") or "").strip().upper()
        msg = _strip_html((d or {}).get("msg", ""))
        if tipo == "WARN" and msg:
            return msg

    return "-"


def _fila_reporte_xlsx(resultado):
    promo_info = resultado.get("promo_info") or {}
    productos_condicion, detalle_condicion = _extraer_productos_y_detalle(resultado.get("resumen_condicion", "-"))
    productos_aplicador, detalle_aplicador = _extraer_productos_y_detalle(resultado.get("resumen_aplicador", "-"))

    fecha_inicio = str(promo_info.get("startDate") or promo_info.get("__start_date") or "-").strip() or "-"
    fecha_fin = str(promo_info.get("endDate") or promo_info.get("__end_date") or "-").strip() or "-"

    return [
        _estado_reporte_descarga(resultado),
        _observacion_reporte(resultado),
        resultado.get("id_geo", "-"),
        fecha_inicio,
        fecha_fin,
        resultado.get("tipo_promocion", "-"),
        promo_info.get("__area_responsable", promo_info.get("area_responsable", "-")),
        _texto_competencia_prolijo(promo_info.get("__tipo_competencia", "-")),
        productos_condicion,
        detalle_condicion,
        productos_aplicador,
        detalle_aplicador,
        promo_info.get("creationUser", "-"),
        resultado.get("excel_origen", "-"),
        resultado.get("export_origen", "-"),
    ]


def _armar_xlsx_resultados(rc, resultados_tradicional, resultados_completar):
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"

    headers = [
        "Estado",
        "Observaciones",
        "ID GEO",
        "Fecha inicio",
        "Fecha fin",
        "Tipo descuento",
        "Área responsable",
        "Tipo competencia",
        "Productos condición",
        "Detalle condición",
        "Productos aplicador",
        "Detalle aplicador",
        "Usuario creador",
        "Excel",
        "Export",
    ]

    title_fill = PatternFill("solid", fgColor="1F4F82")
    title_font = Font(color="FFFFFF", bold=True)
    header_fill = PatternFill("solid", fgColor="DCE6F1")
    header_font = Font(bold=True, color="1F1F1F")
    thin_gray = Side(style="thin", color="D9E0E7")
    border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)

    ws.merge_cells("A1:O1")
    ws["A1"] = "REPORTE DE VALIDACIÓN DE PROMOCIONES"
    ws["A1"].fill = title_fill
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    ws["A2"] = "Fecha validación"
    ws["B2"] = datetime.now().strftime("%d-%m-%Y")
    ws["D2"] = "Usuario filtrado"
    ws["E2"] = rc or "-"

    for cell in ("A2", "D2"):
        ws[cell].font = Font(bold=True, color="44515E")

    header_row = 4
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    data_rows = []
    for r in resultados_tradicional:
        data_rows.append(_fila_reporte_xlsx(r))
    for r in resultados_completar:
        data_rows.append(_fila_reporte_xlsx(r))

    start_row = header_row + 1
    for row_idx, row_data in enumerate(data_rows, start=start_row):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        estado = str(row_data[0])
        estado_cell = ws.cell(row=row_idx, column=1)
        estado_cell.font = Font(bold=True)
        if estado == "OK":
            estado_cell.fill = PatternFill("solid", fgColor="DCEFE5")
            estado_cell.font = Font(bold=True, color="115C39")
        elif estado == "ATENCION":
            estado_cell.fill = PatternFill("solid", fgColor="FFF2C9")
            estado_cell.font = Font(bold=True, color="7A5B00")
        else:
            estado_cell.fill = PatternFill("solid", fgColor="F7D9DC")
            estado_cell.font = Font(bold=True, color="8D2430")

    widths = {
        "A": 12, "B": 42, "C": 12, "D": 14, "E": 14,
        "F": 22, "G": 20, "H": 22, "I": 30, "J": 34,
        "K": 30, "L": 34, "M": 18, "N": 38, "O": 26,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:O{max(4, ws.max_row)}"
    ws.sheet_view.showGridLines = False

    for row in range(5, ws.max_row + 1):
        ws.row_dimensions[row].height = 34

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def serializar_resultados(resultados):

    limpios = []
    for r in resultados:
        copia = dict(r)
        copia["promo_info"] = dict(copia.get("promo_info") or {})
        limpios.append(copia)
    return json.dumps(limpios, ensure_ascii=False)


# ============================================================
# RUTA PRINCIPAL
# ============================================================
@app.route("/")
@app.route("/validPromotion/")
def inicio():
    excel, export = listar_archivos()
    return render_template("index.html", excel_files=excel, export_files=export)



# ============================================================
# SUBIR ARCHIVOS
# ============================================================
@app.route("/upload", methods=["POST"])
def upload_files():
    cargados_excel = 0
    cargados_export = 0
    export_rechazados = []

    FIRMA_EXPORT = (
        "<?xml version='1.0' encoding='UTF-8'?>"
        "<uy.com.geocom.geopromotion.service.promotion.PromotionBlockList>"
        "<promotionTypeList>"
    )

    for file in request.files.getlist("excel_files"):
        if file and file.filename.lower().endswith(".xlsx"):
            file.save(os.path.join(EXCEL_PATH, file.filename))
            cargados_excel += 1

    for file in request.files.getlist("export_files"):
        if not file or not file.filename.lower().endswith(".txt"):
            continue
        try:
            contenido = file.stream.read(4096)
            file.stream.seek(0)
            texto = None
            for enc in ("utf-8", "utf-16", "latin-1"):
                try:
                    texto = contenido.decode(enc)
                    break
                except UnicodeDecodeError:
                    continue
            if texto is None or FIRMA_EXPORT not in texto.replace("\n", "").replace("\r", ""):
                export_rechazados.append(file.filename)
                continue
            file.save(os.path.join(EXPORT_PATH, file.filename))
            cargados_export += 1
        except Exception:
            export_rechazados.append(file.filename)

    return jsonify({
        "mensaje": "Carga finalizada",
        "excel": cargados_excel,
        "export": cargados_export,
        "excel_cargados": cargados_excel,
        "export_validos": cargados_export,
        "export_rechazados": export_rechazados,
        "lista_excel": os.listdir(EXCEL_PATH),
        "lista_export": os.listdir(EXPORT_PATH),
    })


# ============================================================
# BORRAR ARCHIVOS
# ============================================================
@app.route("/borrar", methods=["POST"])
def borrar_archivos():
    tipo = request.form.get("tipo")
    if tipo == "excel":
        errores = limpiar_carpeta(EXCEL_PATH)
        msg = "Se borraron TODOS los Excel."
    elif tipo == "export":
        errores = limpiar_carpeta(EXPORT_PATH)
        msg = "Se borraron TODOS los Export."
    else:
        return jsonify({"error": "Tipo inválido"})
    return jsonify({
        "mensaje": msg,
        "errores": errores,
        "lista_excel": os.listdir(EXCEL_PATH),
        "lista_export": os.listdir(EXPORT_PATH),
    })


def construir_indices_export(export_files):
    promo_info_por_id = {}
    promos_por_id = {}
    listas_productos_export = {}

    for exp in export_files:
        export_name = os.path.basename(exp)
        tree, raw_text = convertir_txt_a_xml_con_root(exp)
        promos = parsear_promos(tree, export_name=export_name)

        try:
            listas_tmp = parsear_listas_productos_export(exp)
            for nombre, productos in listas_tmp.items():
                listas_productos_export.setdefault(nombre, set()).update(productos)
        except Exception:
            pass

        for promo_dict in promos:
            pid = normalizar_local(str(promo_dict.get("id")).split(".")[0])
            promo_info_por_id[pid] = {
                "creationUser": promo_dict.get("creationUser", "-"),
                "enabled": promo_dict.get("enabled", False),
                "__tipo_competencia": promo_dict.get("__tipo_competencia", "-"),
                "__area_responsable": promo_dict.get("area_name", "-"),
                "__export_origen": promo_dict.get("__export_origen", export_name),
                "__tipo_descuento": "-",
                "startDate": promo_dict.get("startDate", "-"),
                "endDate": promo_dict.get("endDate", "-"),
            }
            if pid not in promos_por_id:
                promos_por_id[pid] = promo_dict

    return promo_info_por_id, promos_por_id, listas_productos_export


def construir_mapa_area_responsable(excel_files):
    mapa = {}
    for file in excel_files:
        try:
            df_imput = leer_hoja_imput(file)
        except Exception:
            df_imput = None
        if df_imput is None or df_imput.empty:
            continue
        cols = {str(c).strip().upper(): c for c in df_imput.columns}
        col_id = None
        for k, c in cols.items():
            if "GEOCOM" in k or k in {"ID GEO", "ID GEOCOM", "ID"}:
                col_id = c
                break
        col_area = None
        for k, c in cols.items():
            if "AREARESPONSABLE" in k or "AREA RESPONSABLE" in k or k == "AREA":
                col_area = c
                break
        if not col_id or not col_area:
            continue
        for _, row in df_imput.iterrows():
            pid = normalizar_local(row.get(col_id))
            area = normalizar_texto(row.get(col_area))
            if pid and area:
                mapa[pid] = area
    return mapa


# ============================================================
# PROCESAR VALIDACIÓN
# ============================================================
@app.route("/procesar", methods=["POST"])
def procesar():
    rc_web = request.form.get("rc", "").strip().upper()
    excel_files = [os.path.join(EXCEL_PATH, f) for f in os.listdir(EXCEL_PATH) if f.lower().endswith(".xlsx")]
    export_files = [os.path.join(EXPORT_PATH, f) for f in os.listdir(EXPORT_PATH) if f.lower().endswith(".txt")]

    resultados_tradicional = []
    resultados_completar = []

    promo_info_por_id, promos_por_id, listas_productos_export = construir_indices_export(export_files)
    mapa_area_responsable = construir_mapa_area_responsable(excel_files)

    # FLUJO TRADICIONAL
    if rc_web:
        try:
            df_usuario, _, _, archivos_tradicional = ejecutar_flujo_tradicional(excel_files, rc_externo=rc_web)
        except Exception:
            df_usuario, archivos_tradicional = None, []

        if df_usuario is not None and not df_usuario.empty:
            excel_origen_trad = ", ".join(sorted({os.path.basename(f) for f in archivos_tradicional}))
            for id_geo, grupo in df_usuario.groupby("ID GEO"):
                id_geo_norm = normalizar_local(str(id_geo).split(".")[0])
                promo = promos_por_id.get(id_geo_norm)
                info = promo_info_por_id.get(id_geo_norm, {}).copy()
                if "DESCUENTO" in grupo.columns:
                    val = grupo["DESCUENTO"].iloc[0]
                    if isinstance(val, (int, float)):
                        info["__tipo_descuento"] = f"PORCENTUAL - {int(val * 100) if val <= 1 else int(val)}%"
                    else:
                        info["__tipo_descuento"] = f"PORCENTUAL - {str(val).strip()}"
                else:
                    info["__tipo_descuento"] = "-"

                if promo is None:
                    analisis = {
                        "mensaje_principal": "No existe en export", "aviso_principal": "", "estado_id": "No coinciden",
                        "estado_facturar": "No evaluado", "estado_fechas": "No evaluado", "estado_condicion": "No evaluado",
                        "estado_applier": "No evaluado", "fecha_inicio_ok": None, "fecha_fin_ok": None,
                        "tipo_promocion": "-", "resumen_condicion": "-", "resumen_aplicador": "-",
                    }
                    resultados_tradicional.append(construir_resultado_web(id_geo, excel_origen_trad, "-", info, [{"tipo": "ERR", "msg": "No encontrada en export"}], analisis))
                    continue

                _, detalles = validar_promocion_tradicional(id_geo, grupo, promo, {}, {}, {})
                analisis = analizar_detalles(detalles)
                info["__area_responsable"] = analisis.get("area_responsable", info.get("__area_responsable", "-"))
                resultados_tradicional.append(construir_resultado_web(id_geo, excel_origen_trad, info.get("__export_origen", "-"), info, detalles, analisis))

    # FLUJO COMPLETAR
    df_completar_total = pd.DataFrame()
    for file in excel_files:
        df_c = leer_hoja_completar(file)
        if df_c is not None and not df_c.empty:
            df_c["__excel_origen"] = os.path.basename(file)
            df_completar_total = pd.concat([df_completar_total, df_c], ignore_index=True)

    if not df_completar_total.empty:
        col_id_geo = [c for c in df_completar_total.columns if "GEOCOM" in str(c).upper()][0]
        for id_geo, grupo in df_completar_total.groupby(col_id_geo):
            id_geo_norm = normalizar_local(str(id_geo).split(".")[0])
            promo = promos_por_id.get(id_geo_norm)
            info = promo_info_por_id.get(id_geo_norm, {}).copy()
            excel_origen = grupo["__excel_origen"].iloc[0]

            if promo is None:
                analisis = {
                    "mensaje_principal": "No existe en export", "aviso_principal": "", "estado_id": "No coinciden",
                    "estado_facturar": "No evaluado", "estado_fechas": "No evaluado", "estado_condicion": "No evaluado",
                    "estado_applier": "No evaluado", "fecha_inicio_ok": None, "fecha_fin_ok": None,
                    "tipo_promocion": "-", "resumen_condicion": "-", "resumen_aplicador": "-",
                }
                resultados_completar.append(construir_resultado_web(id_geo, excel_origen, "-", {}, [{"tipo": "ERR", "msg": "No encontrada en export"}], analisis))
                continue

            _, detalles, msje_popup = validar_promocion_completar(
                id_geo,
                grupo,
                promo,
                listas_productos_export,
                mapa_area_responsable=mapa_area_responsable,
                promos_por_id=promos_por_id,
                retornar_msje_data=True,
            )
            analisis = analizar_detalles(detalles)
            info["__tipo_descuento"] = analisis["tipo_promocion"] or promo.get("__tipo_descuento", "-")
            info["__area_responsable"] = analisis.get("area_responsable", info.get("__area_responsable", "-"))

            resultado_principal = construir_resultado_web(
                id_geo,
                excel_origen,
                info.get("__export_origen", "-"),
                info,
                detalles,
                analisis,
                msje_popup=msje_popup,
            )
            resultados_completar.append(resultado_principal)

    todos_los_resultados = resultados_tradicional + resultados_completar
    total_promos = len(todos_los_resultados)
    total_ok = sum(1 for r in todos_los_resultados if r.get("mensaje") == "Coinciden" and not r.get("aviso_principal", "").startswith("Posible extensión"))
    total_warn = sum(1 for r in todos_los_resultados if r.get("aviso_principal"))
    total_err = sum(1 for r in todos_los_resultados if r.get("mensaje") != "Coinciden")

    global ULTIMO_REPORTE_DESCARGA
    ULTIMO_REPORTE_DESCARGA = {
        "rc": rc_web,
        "tradicional": _copiar_resultados_para_descarga(resultados_tradicional),
        "completar": _copiar_resultados_para_descarga(resultados_completar),
    }

    return render_template(
        "resultado.html",
        rc=rc_web,
        resultados_tradicional=resultados_tradicional,
        resultados_completar=resultados_completar,
        total_promos=total_promos,
        total_ok=total_ok,
        total_warn=total_warn,
        total_err=total_err,
        tradicional_data=serializar_resultados(resultados_tradicional),
        completar_data=serializar_resultados(resultados_completar),
    )


# ============================================================
# DESCARGAR RESULTADOS
# ============================================================
@app.route("/descargar_resultados", methods=["POST"])
def descargar_resultados():
    global ULTIMO_REPORTE_DESCARGA

    rc = (ULTIMO_REPORTE_DESCARGA or {}).get("rc", "")
    resultados_tradicional = list((ULTIMO_REPORTE_DESCARGA or {}).get("tradicional", []))
    resultados_completar = list((ULTIMO_REPORTE_DESCARGA or {}).get("completar", []))

    if not resultados_tradicional and not resultados_completar:
        salida = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte"
        ws["A1"] = "No hay resultados disponibles para descargar."
        ws["A2"] = "Primero ejecuta una validación."
        wb.save(salida)
        salida.seek(0)
        return send_file(
            salida,
            as_attachment=True,
            download_name=f"resultado_validacion_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    salida = _armar_xlsx_resultados(rc, resultados_tradicional, resultados_completar)
    return send_file(
        salida,
        as_attachment=True,
        download_name=f"resultado_validacion_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    app.run(debug=True)
